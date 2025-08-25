"""
Utilities to update Excel reports with scoring metrics
"""

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from typing import Dict

def update_excel_with_scoring(excel_path: str, scoring_results: Dict, scoring_summary: pd.DataFrame):
    """Update existing Excel report with scoring results"""
    
    try:
        # Load existing workbook
        workbook = load_workbook(excel_path)
        
        # Create scoring summary sheet
        if 'scoring_summary' in workbook.sheetnames:
            del workbook['scoring_summary']
        
        ws_summary = workbook.create_sheet('scoring_summary')
        
        # Add summary data
        ws_summary['A1'] = 'Scoring Summary Report'
        ws_summary['A1'].font = ws_summary['A1'].font.copy(bold=True, size=14)
        
        # Write summary table
        for idx, row in scoring_summary.iterrows():
            ws_summary.cell(row=idx+3, column=1, value=row['Metric'])
            ws_summary.cell(row=idx+3, column=2, value=row['Value'])
        
        # Add headers
        ws_summary['A2'] = 'Metric'
        ws_summary['B2'] = 'Value'
        ws_summary['A2'].font = ws_summary['A2'].font.copy(bold=True)
        ws_summary['B2'].font = ws_summary['B2'].font.copy(bold=True)
        
        # Auto-adjust column widths
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        # Create score distribution sheet
        if 'score_distribution' in workbook.sheetnames:
            del workbook['score_distribution']
            
        ws_dist = workbook.create_sheet('score_distribution')
        
        # Score statistics
        scores = scoring_results['scores']
        score_stats = pd.DataFrame({
            'Statistic': ['Count', 'Mean', 'Std', 'Min', '25%', '50%', '75%', 'Max'],
            'Value': [
                len(scores),
                scores.mean(),
                scores.std(),
                scores.min(),
                pd.Series(scores).quantile(0.25),
                pd.Series(scores).quantile(0.50),
                pd.Series(scores).quantile(0.75),
                scores.max()
            ]
        })
        
        ws_dist['A1'] = 'Score Distribution Statistics'
        ws_dist['A1'].font = ws_dist['A1'].font.copy(bold=True, size=14)
        
        ws_dist['A2'] = 'Statistic'
        ws_dist['B2'] = 'Value'
        ws_dist['A2'].font = ws_dist['A2'].font.copy(bold=True)
        ws_dist['B2'].font = ws_dist['B2'].font.copy(bold=True)
        
        for idx, row in score_stats.iterrows():
            ws_dist.cell(row=idx+3, column=1, value=row['Statistic'])
            ws_dist.cell(row=idx+3, column=2, value=f"{row['Value']:.4f}" if isinstance(row['Value'], float) else str(row['Value']))
        
        ws_dist.column_dimensions['A'].width = 15
        ws_dist.column_dimensions['B'].width = 15
        
        # Save the updated workbook
        workbook.save(excel_path)
        workbook.close()
        
        print(f"✅ Excel report updated with scoring metrics: {excel_path}")
        print("   - Added 'scoring_summary' sheet")
        print("   - Added 'score_distribution' sheet")
        
        return True
        
    except Exception as e:
        print(f"❌ Error updating Excel report: {e}")
        return False

def create_comprehensive_report(pipeline_results: Dict, scoring_results: Dict, output_path: str):
    """Create a comprehensive Excel report with both pipeline and scoring results"""
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Pipeline summary
        pipeline_summary = pd.DataFrame({
            'Metric': [
                'Best Model',
                'Final Features Count',
                'Run ID',
                'Training Records',
                'Test Records',
                'OOT Records'
            ],
            'Value': [
                pipeline_results.get('best_model', 'Unknown'),
                len(pipeline_results.get('final_features', [])),
                pipeline_results.get('run_id', 'Unknown'),
                'N/A',  # Would need to be passed from pipeline
                'N/A',
                'N/A'
            ]
        })
        
        pipeline_summary.to_excel(writer, sheet_name='pipeline_summary', index=False)
        
        # Scoring summary
        scoring_summary = pd.DataFrame({
            'Metric': [
                'Total Scored Records',
                'Records with Target',
                'Records without Target',
                'AUC (with targets)',
                'Gini Coefficient',
                'KS Statistic',
                'Default Rate',
                'PSI Score'
            ],
            'Value': [
                scoring_results['n_total'],
                scoring_results['n_with_target'],
                scoring_results['n_without_target'],
                f"{scoring_results.get('auc', 0):.4f}",
                f"{scoring_results.get('gini', 0):.4f}",
                f"{scoring_results.get('ks', 0):.4f}",
                f"{scoring_results.get('default_rate', 0):.3f}",
                f"{scoring_results.get('psi_score', 0):.4f}" if scoring_results.get('psi_score') is not None else 'N/A'
            ]
        })
        
        scoring_summary.to_excel(writer, sheet_name='scoring_summary', index=False)
        
        # Score distribution
        scores = scoring_results['scores']
        score_stats = pd.DataFrame({
            'Statistic': ['Count', 'Mean', 'Std', 'Min', '25%', '50%', '75%', 'Max'],
            'Value': [
                len(scores),
                f"{scores.mean():.4f}",
                f"{scores.std():.4f}",
                f"{scores.min():.4f}",
                f"{pd.Series(scores).quantile(0.25):.4f}",
                f"{pd.Series(scores).quantile(0.50):.4f}",
                f"{pd.Series(scores).quantile(0.75):.4f}",
                f"{scores.max():.4f}"
            ]
        })
        
        score_stats.to_excel(writer, sheet_name='score_distribution', index=False)
        
    print(f"✅ Comprehensive report created: {output_path}")
    
    return True